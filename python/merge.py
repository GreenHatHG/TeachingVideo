# -*- coding: UTF-8 -*-

import openpyxl

#对搜索到的结果进行过滤并且合并到一个sheet
def merge(keyword, file):
    #名称列表
    titles = []
    #链接列表
    links = []
    #将关键字转为小写，方便查找
    keyword = keyword.lower()

    try:
        wb = openpyxl.load_workbook(file)
        sheets = wb.sheetnames

        #清空excel历史数据
        if '合并' in str(sheets):
            wb.remove(wb['合并'])
        wb.create_sheet('合并')
        ws = wb['合并']
        ws.append(('名称', '链接'))

        #提取包含关键字的信息到新的sheet
        for sheet in sheets:
            w = wb[sheet]
            row_num = w.max_row
            if sheet == '网易云课堂':
                for i in range(row_num):
                    i += 1
                    row = w[i]
                    if keyword in row[0].value.lower():
                        titles.append(row[0].value)
                        links.append(row[4].value)

            elif sheet == 'Bilibili':
                for i in range(row_num):
                    i += 1
                    row = w[i]
                    if keyword in row[0].value.lower():
                        titles.append(row[0].value)
                        links.append(row[3].value)

        #写入excel
        for i in range(len(titles)):
            ws.append((titles[i], links[i]))
        wb.save(file)
        wb.close()
        print("合并成功")

    except Exception as e:
        print("合并失败")
        print(e)
