# -*- coding: UTF-8 -*-
#爬取网易云课堂主代码

import requests
import time
import json
import openpyxl

#网易云课堂视频信息接口
url = "https://study.163.com/p/search/studycourse.json"

#名称
productName = []
#评分
score = []
#价格
price = []
#链接
link = []


def getContent(keyword, index):
    #请求接口所需要的数据
    payload = {
        "activityId": 0,
        # 关键字
        "keyword": keyword,
        "orderType": 5,
        #第几页
        "pageIndex": index,
        #每一页数据量
        "pageSize": 50,
        "priceType": -1,
        "qualityType": 0,
        "relativeOffset": 0,
        "searchTimeType": -1
    }
    headers = {
        "accept": "application/json",
        "origin": "https://study.163.com",
        "content-type": "application/json",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.132 Safari/537.36"
    }
    try:
        #请求接口并且得到数据进行提取
        res = requests.post(url, json=payload, headers=headers)
        content_json = json.loads(res.text)
        if content_json['result']['query'] is not None and content_json['code'] == 0:
            data = getList(content_json)
            if data is not None:
                for item in data:
                    if len(item) > 0:
                        print(item["productName"])
                        productName.append(item["productName"])
                        score.append(item["score"])
                        price.append(
                            item['discountPrice'] if item['discountPrice'] is not None else item['originalPrice'])
                        link.append("https://study.163.com/course/introduction/" + str(item['productId']) + ".htm")
            return True
        else:
            return False
    except Exception as e:
        print("error")
        print(e)


def getList(content_json):
    list = content_json['result']['list']
    if (list is not None):
        return list
    else:
        return None

#保存信息到excel
def save_to_excel(file):
    global wb
    try:
        wb = openpyxl.load_workbook(file)

        # 清空excel历史数据
        if '网易云课堂' in str(wb.sheetnames):
            wb.remove(wb['网易云课堂'])
        wb.create_sheet('网易云课堂')
        ws = wb['网易云课堂']

        #写入excel
        ws.append(('名称', '评分', '价格', '链接'))
        for i in range(len(productName)):
            ws.append((productName[i], score[i], price[i], link[i]))
        wb.save(file)
        wb.close()
        print('写入excel成功')
    except Exception as e:
        print('写入excel失败，只写入部分数据，可能是因为写入的数据量过大')
        wb.save(file)
        wb.close()
        print(e)


def start(keyword, file):
    for i in range(1, 100):
        if not getContent(keyword, i):
            break
        time.sleep(2)
    save_to_excel(file)
