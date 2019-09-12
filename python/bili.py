# -*- coding: UTF-8 -*-
# 爬取bili主代码

import urllib.parse
import re
import requests
import openpyxl
import time
from bs4 import BeautifulSoup

# 名称
titles = []
# 播放量
watch_nums = []
# 上传时间
times = []
# 链接
links = []


#得到网页源码
def getText(url):
    try:
        kv = {'user-agent': 'Mozilla/5.0'}
        r = requests.get(url, headers=kv)
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        return r.text
    except:
        return '错误'

#处理网页源码
def getContext(text):
    global titles, watch_nums, times, links
    soup = BeautifulSoup(text, 'lxml')
    li = soup.select(".mixin-list > ul > li:nth-child(n)")
    if len(li) == 0:
        li = soup.select(".flow-loader > ul > li:nth-child(n)")

    if 0 >= len(li) or li is None:
        return False
    else:
        for item in li:
            title = item.select("div > div.headline.clearfix > a")
            print(title[0]['title'])
            titles.append(title[0]['title'])

            watch_num = item.select("div > div.tags > span.so-icon.watch-num")
            watch_nums.append(watch_num[0].text.strip())

            time = item.select("div > div.tags > span.so-icon.time")
            times.append(time[0].text.strip())

            linkTmp = item.select("li > a")
            link = re.findall(r'//(.+)\?from=search', str(linkTmp))
            links.append(link[0])

        return True

#保存到excel
def save_to_excel(file):
    try:
        wb = openpyxl.load_workbook(file)

        # 清空excel历史数据
        if 'Bilibili' in str(wb.sheetnames):
            wb.remove(wb['Bilibili'])
        wb.create_sheet('Bilibili')
        ws = wb['Bilibili']

        # 写入excel
        ws.append(('名称', '播放量', '上传时间', '链接'))
        for i in range(len(titles)):
            ws.append((titles[i], watch_nums[i], times[i], links[i]))
        wb.save(file)
        wb.close()
        print('写入excel成功')
    except Exception as e:
        print('写入excel失败')
        print(e)


def start(keyword, file):
    #处理除英文外的字符
    keyword = urllib.parse.quote(keyword)
    pattern = "https://search.bilibili.com/all?keyword=" + keyword + "&page="
    for i in range(1, 2):
        url = pattern + str(i)
        text = getText(url)
        if not getContext(text):
            break
        time.sleep(2)
    save_to_excel(file)
